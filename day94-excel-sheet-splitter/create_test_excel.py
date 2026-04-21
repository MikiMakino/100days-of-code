"""
テスト用Excelブックを生成するスクリプト。
ex.py の動作確認用。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT_FILE = "test_book.xlsx"

# シート構成
COMMON_SHEETS = ["共通シート"]          # 全ファイルに含める想定
EXCLUDE_SHEETS = ["除外シート_マスタ"]  # どのファイルにも含めない想定
SPLIT_SHEETS = [                         # 1枚ずつ分割される想定
    "A社",
    "B社",
    "C社",
    "D社",
]

COLORS = {
    "共通シート":       "4472C4",  # 青
    "除外シート_マスタ": "FF0000",  # 赤
    "A社":      "70AD47",  # 緑
    "B社":      "ED7D31",  # オレンジ
    "C社":      "FFC000",  # 黄
    "D社":      "9B59B6",  # 紫
}

wb = Workbook()
wb.remove(wb.active)  # デフォルトの空シートを削除

for sheet_name in COMMON_SHEETS + EXCLUDE_SHEETS + SPLIT_SHEETS:
    ws = wb.create_sheet(title=sheet_name)
    color = COLORS.get(sheet_name, "CCCCCC")

    # ヘッダー行
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30

    title_cell = ws["A1"]
    title_cell.value = f"シート: {sheet_name}"
    title_cell.font = Font(bold=True, color="FFFFFF", size=14)
    title_cell.fill = PatternFill("solid", fgColor=color)
    title_cell.alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    ws["A2"] = "項目"
    ws["B2"] = "内容"
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    ws["A3"] = "シート種別"
    if sheet_name in COMMON_SHEETS:
        ws["B3"] = "共通シート（全ファイルに含める）"
    elif sheet_name in EXCLUDE_SHEETS:
        ws["B3"] = "除外シート（どのファイルにも含めない）"
    else:
        ws["B3"] = "分割対象シート（個別ファイルに分ける）"

    ws["A4"] = "サンプルデータ"
    ws["B4"] = f"{sheet_name} のデータ行1"
    ws["A5"] = ""
    ws["B5"] = f"{sheet_name} のデータ行2"

wb.save(OUT_FILE)
print(f"作成完了: {OUT_FILE}")
print(f"  シート構成 ({len(wb.sheetnames)}枚): {wb.sheetnames}")
print()
print("【ex.py での実行例】")
print(f"  python ex.py {OUT_FILE}")
print(f"  -> 共通シート選択 : 「共通シート」")
print(f"  -> 除外シート選択 : 「除外シート_マスタ」")
print(f"  -> 分割対象       : A〜D社 の4ファイルが出力されます")
