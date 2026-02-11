"""
Day 87 - Client Master Import
===============================
取引先マスタExcelからSQLiteにインポートする。

Usage:
    python client_import.py                     # インポート実行
    python client_import.py --file other.xlsx   # 別ファイル指定
    python client_import.py --list              # 登録済み取引先一覧
"""
import sys
import os
import logging
import argparse

from openpyxl import load_workbook

from config import (
    CLIENT_MASTER_EXCEL, CLIENT_SHEET_NAME,
    CLIENT_COL_CODE, CLIENT_COL_NAME,
    CLIENT_COL_MONTH1, CLIENT_COL_MONTH2, CLIENT_COL_CS,
    LOG_FILE, LOG_LEVEL, LOG_FORMAT, LOG_DATE_FORMAT,
)
from database import init_db, upsert_client, get_all_clients

logger = logging.getLogger(__name__)


def setup_logging():
    """ログ設定。"""
    root = logging.getLogger()
    root.setLevel(getattr(logging, LOG_LEVEL))
    if root.handlers:
        return
    fmt = logging.Formatter(LOG_FORMAT, datefmt=LOG_DATE_FORMAT)
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    root.addHandler(fh)
    root.addHandler(ch)


def import_from_excel(filepath: str) -> int:
    """
    Excelファイルから取引先マスタをインポートする。
    既存コードは上書き更新（UPSERT）。
    Returns: インポート件数
    """
    if not os.path.exists(filepath):
        logger.error("ファイルが見つかりません: %s", filepath)
        return 0

    wb = load_workbook(filepath, read_only=True, data_only=True)

    if CLIENT_SHEET_NAME not in wb.sheetnames:
        logger.error("シート '%s' が見つかりません。存在するシート: %s",
                      CLIENT_SHEET_NAME, wb.sheetnames)
        wb.close()
        return 0

    ws = wb[CLIENT_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        logger.warning("シートが空です")
        return 0

    # --- ヘッダー行から列位置を特定 ---
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_map = {}
    required = {
        CLIENT_COL_CODE: "code",
        CLIENT_COL_NAME: "name",
    }
    optional = {
        CLIENT_COL_MONTH1: "month_1",
        CLIENT_COL_MONTH2: "month_2",
        CLIENT_COL_CS: "cs_category",
    }

    for col_name, key in {**required, **optional}.items():
        if col_name in header:
            col_map[key] = header.index(col_name)
        elif key in required.values():
            logger.error("必須列 '%s' がヘッダーに見つかりません。ヘッダー: %s",
                          col_name, header)
            return 0

    logger.info("列マッピング: %s", col_map)

    # --- データ行を処理 ---
    count = 0
    for i, row in enumerate(rows[1:], start=2):
        code_val = row[col_map["code"]] if col_map.get("code") is not None else None
        name_val = row[col_map["name"]] if col_map.get("name") is not None else None

        if not code_val or not name_val:
            logger.debug("行%d: コードまたは社名が空のためスキップ", i)
            continue

        code = str(code_val).strip()
        name = str(name_val).strip()
        m1 = str(row[col_map["month_1"]]).strip() if col_map.get("month_1") is not None and row[col_map["month_1"]] else ""
        m2 = str(row[col_map["month_2"]]).strip() if col_map.get("month_2") is not None and row[col_map["month_2"]] else ""
        cs = str(row[col_map["cs_category"]]).strip() if col_map.get("cs_category") is not None and row[col_map["cs_category"]] else ""

        upsert_client(code, name, m1, m2, cs)
        count += 1
        logger.debug("UPSERT: [%s] %s", code, name)

    logger.info("インポート完了: %d件", count)
    return count


def list_clients():
    """登録済み取引先を一覧表示する。"""
    clients = get_all_clients()
    if not clients:
        print("取引先が登録されていません。")
        return

    print(f"\n{'コード':<10} {'社名':<30} {'実施月1':<8} {'実施月2':<8} {'CS区分':<10}")
    print("-" * 70)
    for c in clients:
        print(f"{c['code']:<10} {c['company_name']:<30} "
              f"{c['month_1'] or '':<8} {c['month_2'] or '':<8} "
              f"{c['cs_category'] or '':<10}")
    print(f"\n合計: {len(clients)}件")


def main():
    parser = argparse.ArgumentParser(
        description="取引先マスタ Excel → SQLite インポート"
    )
    parser.add_argument(
        "--file", default=CLIENT_MASTER_EXCEL,
        help=f"取引先マスタExcelのパス (デフォルト: {CLIENT_MASTER_EXCEL})",
    )
    parser.add_argument(
        "--list", action="store_true",
        help="登録済み取引先を一覧表示",
    )
    args = parser.parse_args()

    setup_logging()
    init_db()

    if args.list:
        list_clients()
    else:
        count = import_from_excel(args.file)
        if count > 0:
            print(f"\n{count}件の取引先をインポートしました。")
            list_clients()
        else:
            print("\nインポートされた取引先はありません。")


if __name__ == "__main__":
    main()
