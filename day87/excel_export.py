"""
Day 87 - Excel Export
======================
管理台帳 + イベントログ + 取引先マスタ + ファイル処理結果を
4シート構成のExcelファイルにエクスポートする。
"""
import os
import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import (
    EXPORT_DIR, EXCEL_FILENAME_TEMPLATE,
    EXCEL_SHEET_NAME, EXCEL_EVENTS_SHEET_NAME, EXCEL_CLIENTS_SHEET_NAME,
    EXCEL_FILE_RESULTS_SHEET_NAME,
)
from database import (
    fetch_all_records, fetch_events, get_all_clients, fetch_file_results,
)

logger = logging.getLogger(__name__)

# --- 台帳シートの定義 ---
LEDGER_HEADERS = [
    "ID", "取引先コード", "社名", "実施月1", "実施月2", "CS区分",
    "メールボックス", "フォルダ", "送信者", "送信者SMTP",
    "返信先", "To", "Cc", "件名",
    "受信日時", "添付数", "添付ファイル名", "添付種類",
    "保存先フォルダ", "URL数", "処理日時",
]
LEDGER_WIDTHS = [
    6, 12, 22, 8, 8, 10,
    18, 14, 22, 30,
    25, 30, 30, 40,
    20, 8, 40, 20,
    40, 8, 20,
]

# --- イベントログシートの定義 ---
EVENT_HEADERS = [
    "ID", "メッセージID(先頭20)", "日時", "イベント種別",
    "レベル", "ソース", "詳細",
]
EVENT_WIDTHS = [6, 22, 20, 22, 8, 18, 60]

# --- 取引先マスタシートの定義 ---
CLIENT_HEADERS = [
    "コード", "社名", "実施月1", "実施月2", "CS区分",
    "登録日", "更新日",
]
CLIENT_WIDTHS = [12, 30, 10, 10, 12, 20, 20]

# --- ファイル処理結果シートの定義 ---
FILE_RESULT_HEADERS = [
    "ID", "取引先コード", "社名", "ファイル名", "ファイル種別",
    "PDF種別", "件数", "フォルダ", "処理日時",
]
FILE_RESULT_WIDTHS = [6, 12, 22, 40, 10, 10, 8, 40, 20]


def _apply_header_style(ws, headers, widths):
    """ヘッダー行にスタイルを適用する。"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for col_idx, width in enumerate(widths, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width

    return thin_border


def _write_ledger_sheet(wb):
    """台帳シートを書き込む（取引先情報をJOIN済み）。"""
    records = fetch_all_records()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME

    thin_border = _apply_header_style(ws, LEDGER_HEADERS, LEDGER_WIDTHS)

    for row_idx, r in enumerate(records, 2):
        row_data = [
            r["id"], r["client_code"] or "",
            r["company_name"] or "", r["month_1"] or "",
            r["month_2"] or "", r["cs_category"] or "",
            r["mailbox"], r["folder_path"],
            r["sender"], r["sender_smtp"],
            r["reply_to"], r["to_recipients"], r["cc_recipients"],
            r["subject"], r["received_date"],
            r["attachment_count"], r["attachment_names"],
            r["attachment_types"], r["save_folder"],
            r["url_count"], r["processed_at"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return len(records)


def _write_events_sheet(wb):
    """イベントログシートを書き込む。"""
    events = fetch_events()
    ws = wb.create_sheet(title=EXCEL_EVENTS_SHEET_NAME)

    thin_border = _apply_header_style(ws, EVENT_HEADERS, EVENT_WIDTHS)

    error_fill = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    warn_fill = PatternFill(
        start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
    )

    for row_idx, ev in enumerate(events, 2):
        msg_id_short = (ev["message_id"] or "")[:20]
        row_data = [
            ev["id"], msg_id_short, ev["event_time"],
            ev["event_type"], ev["level"], ev["source"],
            ev["detail"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if ev["level"] == "ERROR":
                cell.fill = error_fill
            elif ev["level"] == "WARN":
                cell.fill = warn_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return len(events)


def _write_clients_sheet(wb):
    """取引先マスタシートを書き込む。"""
    clients = get_all_clients()
    ws = wb.create_sheet(title=EXCEL_CLIENTS_SHEET_NAME)

    thin_border = _apply_header_style(ws, CLIENT_HEADERS, CLIENT_WIDTHS)

    for row_idx, c in enumerate(clients, 2):
        row_data = [
            c["code"], c["company_name"],
            c["month_1"] or "", c["month_2"] or "",
            c["cs_category"] or "",
            c["created_at"], c["updated_at"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return len(clients)


def _write_file_results_sheet(wb):
    """ファイル処理結果シートを書き込む。"""
    results = fetch_file_results()
    ws = wb.create_sheet(title=EXCEL_FILE_RESULTS_SHEET_NAME)

    thin_border = _apply_header_style(
        ws, FILE_RESULT_HEADERS, FILE_RESULT_WIDTHS
    )

    for row_idx, r in enumerate(results, 2):
        row_data = [
            r["id"], r["client_code"] or "",
            r["company_name"] or "",
            r["file_name"], r["file_type"],
            r["pdf_type"] or "-",
            r["record_count"], r["source_folder"],
            r["processed_at"],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    return len(results)


def export_to_excel() -> str:
    """
    台帳+イベントログ+取引先マスタ+ファイル処理結果を
    4シート構成のExcelファイルに出力する。
    生成ファイルのパスを返す。レコードがなければ空文字を返す。
    """
    records = fetch_all_records()
    if not records:
        logger.warning("エクスポート対象のレコードがありません")
        return ""

    os.makedirs(EXPORT_DIR, exist_ok=True)

    wb = Workbook()
    record_count = _write_ledger_sheet(wb)
    event_count = _write_events_sheet(wb)
    client_count = _write_clients_sheet(wb)
    file_result_count = _write_file_results_sheet(wb)

    date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = EXCEL_FILENAME_TEMPLATE.format(date=date_str)
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    logger.info("Excelエクスポート完了: %s "
                "(台帳%d件, ログ%d件, 取引先%d件, ファイル処理%d件)",
                filepath, record_count, event_count,
                client_count, file_result_count)
    return filepath
